import io
import re
import pandas as pd
import streamlit as st
import openpyxl

# =========================================================================
# CONFIGURAÇÃO DA PÁGINA
# =========================================================================
st.set_page_config(
    page_title="Consulta de Cargas e Rotas", page_icon="📦", layout="wide"
)

st.title("📦 Consulta Rápida de Cargas por Rota e SKU")
st.markdown(
    "Busque o SKU e a rota para localizar onde o material está e realize a conferência na caixa."
)


def extrair_rota_limpa(valor_rota):
    """Extrai apenas o código da rota (ex: BRGP-BR BR0551285_BRGP -> BR0551285)"""
    if pd.isna(valor_rota):
        return "N/A"
    texto = str(valor_rota).strip()
    match = re.search(r"(BR\d+)", texto, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return texto


def descobrir_linha_cabecalho(caminho_arquivo, nome_aba):
    """Abre o arquivo usando openpyxl nativo e retorna o índice correto do cabeçalho (0-indexed)"""
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True, read_only=True)
    if nome_aba not in wb.sheetnames:
        return 0
    
    ws = wb[nome_aba]
    # Varre as primeiras 15 linhas usando células nativas (sem passar pelo Pandas)
    for idx_linha, linha in enumerate(ws.iter_rows(max_row=15, values_only=True)):
        # Junta todas as células da linha numa única string minúscula
        texto_linha = " ".join([str(celula).strip().lower() for celula in linha if celula is not None])
        if "order number" in texto_linha or "orderkey" in texto_linha:
            return idx_linha
            
    return 0


def carregar_e_limpar_aba(caminho_arquivo, nome_aba):
    """Localiza a linha real do cabeçalho nativamente e carrega o DataFrame limpo"""
    linha_header = descobrir_linha_cabecalho(caminho_arquivo, nome_aba)
    
    # Carrega definindo exatamente qual linha contém os títulos reais das colunas
    df = pd.read_excel(caminho_arquivo, sheet_name=nome_aba, header=linha_header)
    
    # Padroniza nomes de colunas
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    # Normaliza variações para garantir que vire 'ORDERKEY'
    renomear = {}
    for col in df.columns:
        if "ORDER" in col and "NUM" in col:
            renomear[col] = "ORDERKEY"
    if renomear:
        df = df.rename(columns=renomear)
        
    return df


@st.cache_data(ttl=60)
def carregar_dados_local():
    try:
        arquivo_excel = "Export.xlsx"
        
        # Carrega as duas abas aplicando a limpeza nativa de cabeçalho superior
        df_detail = carregar_e_limpar_aba(arquivo_excel, "Detail")
        df_data = carregar_e_limpar_aba(arquivo_excel, "Data")

        # Garante que a chave ORDERKEY seja String idêntica e sem decimais flutuantes
        if "ORDERKEY" in df_detail.columns:
            df_detail["ORDERKEY"] = (
                df_detail["ORDERKEY"]
                .astype(str)
                .str.strip()
                .str.replace(".0", "", regex=False)
            )
        if "ORDERKEY" in df_data.columns:
            df_data["ORDERKEY"] = (
                df_data["ORDERKEY"]
                .astype(str)
                .str.strip()
                .str.replace(".0", "", regex=False)
            )

        # Mapeia dinamicamente a coluna de Rota (ROUTE)
        col_route = [c for c in df_data.columns if "ROUTE" in c]
        if col_route:
            df_data["ROTA_LIMPA"] = df_data[col_route[0]].apply(extrair_rota_limpa)
        else:
            df_data["ROTA_LIMPA"] = "N/A"

        # Mapeia dinamicamente a coluna de Parada (STOP)
        col_stop = [c for c in df_data.columns if "STOP" in c]
        if col_stop:
            df_data["PEDIDO_ROTA"] = (
                df_data[col_stop[0]].astype(str).str.replace(".0", "", regex=False)
            )
        else:
            df_data["PEDIDO_ROTA"] = "N/A"

        # Mapeia dinamicamente SKU e Quantidade
        col_sku = [c for c in df_detail.columns if "SKU" in c]
        col_qty = [c for c in df_detail.columns if "QTY" in c or "QUANT" in c]

        sku_lbl = col_sku[0] if col_sku else "SKU"
        qty_lbl = col_qty[0] if col_qty else "OPENQTY"

        # Filtra apenas o necessário para a tabela final de conferência
        df_det_res = df_detail[["ORDERKEY", sku_lbl, qty_lbl]].rename(
            columns={sku_lbl: "SKU", qty_lbl: "OPENQTY"}
        )
        df_dat_res = df_data[["ORDERKEY", "ROTA_LIMPA", "PEDIDO_ROTA"]]

        # Faz o cruzamento estruturado (PROCV via Left Join)
        df_consolidado = pd.merge(df_det_res, df_dat_res, on="ORDERKEY", how="left")
        return df_consolidado

    except FileNotFoundError:
        st.error(
            "Erro: O arquivo 'Export.xlsx' não foi encontrado na pasta do projeto."
        )
        return None
    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")
        return None


# Botão de controle de cache na barra lateral
if st.sidebar.button("🔄 Forçar Limpeza de Cache"):
    st.cache_data.clear()
    st.rerun()

df_base = carregar_dados_local()

if df_base is not None:
    col1, col2 = st.columns(2)
    with col1:
        sku_busca = st.text_input("🔍 Digite o SKU:", placeholder="Ex: 10226403")
    with col2:
        rota_busca = st.text_input("📍 Digite a Rota:", placeholder="Ex: BR0551285")

    if sku_busca or rota_busca:
        df_filtrado = df_base.copy()

        if sku_busca:
            df_filtrado = df_filtrado[
                df_filtrado["SKU"].astype(str).str.contains(sku_busca, case=False)
            ]
        if rota_busca:
            df_filtrado = df_filtrado[
                df_filtrado["ROTA_LIMPA"]
                .astype(str)
                .str.contains(rota_busca, case=False)
            ]

        if not df_filtrado.empty:
            df_exibicao = pd.DataFrame(
                {
                    "Conferido ✔": [False] * len(df_filtrado),
                    "Ordem (OrderKey)": df_filtrado["ORDERKEY"],
                    "Pedido na Rota": df_filtrado["PEDIDO_ROTA"],
                    "SKU": df_filtrado["SKU"],
                    "Quantia Solicitada": df_filtrado["OPENQTY"],
                    "Rota": df_filtrado["ROTA_LIMPA"],
                }
            )

            st.write(f"### 📋 {len(df_exibicao)} caixas encontradas:")

            st.data_editor(
                df_exibicao,
                hide_index=True,
                use_container_width=True,
                disabled=[
                    "Ordem (OrderKey)",
                    "Pedido na Rota",
                    "SKU",
                    "Quantia Solicitada",
                    "Rota",
                ],
            )
        else:
            st.warning("Nenhum registro encontrado para os filtros aplicados.")
    else:
        st.info(
            "💡 Digite um SKU ou Rota acima para listar as ordens de carregamento e quantias."
        )
else:
    st.info("Aguardando carregamento da base de dados...")
